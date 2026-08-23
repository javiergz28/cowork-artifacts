"""Carga el catálogo/maestro de SKU y audita la consistencia de la convención.

Hallazgo que motivó este módulo (23/08/2026): el mal llamado 'typo de ML' no
viene de ML. El propio catálogo `Productos categorias y enlaces.xlsx` tiene
13 SKU escritos PHC0#### (sin la O) y 40 escritos PHCO####. Cero colisiones
entre ambas formas, así que la corrección es mecánica y segura — pero se venía
haciendo a mano, ciclo a ciclo, sobre SKU distintos cada vez.
"""
import re
import openpyxl
from .util import norm_txt

RE_SKU = re.compile(r"^P[A-Z]{2,4}\d{3,5}[A-Z]*$")


def cargar_catalogo(path, hoja=None, columna=None):
    """Devuelve (set_de_sku, meta). Prefiere una columna declarada ('COD'/'SKU').
    Si no la encuentra, cae al barrido de toda la hoja y lo AVISA — porque un
    barrido puede arrastrar SKU viejos o mal escritos de hojas auxiliares.
    """
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    hojas = [wb[hoja]] if hoja else wb.worksheets
    skus, metodo, usada = set(), "barrido", None
    for ws in hojas:
        filas = list(ws.iter_rows(values_only=True))
        if not filas:
            continue
        hdr = [norm_txt(c) for c in filas[0]]
        objetivo = norm_txt(columna) if columna else None
        idx = None
        for j, h in enumerate(hdr):
            if objetivo and h == objetivo:
                idx = j
                break
            if not objetivo and h in ("cod", "sku", "codigo", "cod."):
                idx = j
                break
        if idx is not None:
            for r in filas[1:]:
                if len(r) > idx and r[idx]:
                    s = re.sub(r"\s+", "", str(r[idx])).upper()
                    if RE_SKU.match(s):
                        skus.add(s)
            metodo, usada = "columna", f"{ws.title}!{filas[0][idx]}"
            break
    if metodo == "barrido":
        for ws in hojas:
            for r in ws.iter_rows(values_only=True):
                for c in r:
                    if c:
                        s = re.sub(r"\s+", "", str(c)).upper()
                        if RE_SKU.match(s):
                            skus.add(s)
    wb.close()
    return skus, {"metodo": metodo, "columna": usada, "total": len(skus)}


def auditar(skus):
    """Detecta inconsistencias de convención en el propio catálogo."""
    sin_o = sorted(s for s in skus if re.match(r"^PHC0\d{4}", s) and not s.startswith("PHCO"))
    colisiones = [s for s in sin_o if "PHCO" + s[4:] in skus]
    largos = sorted(s for s in skus if re.match(r"^PHCO\d{5}", s))
    familias = {}
    for s in skus:
        m = re.match(r"^(P[A-Z]{2,4})", s)
        if m:
            familias[m.group(1)] = familias.get(m.group(1), 0) + 1
    return {
        "phc_sin_o": sin_o,
        "colisiones": colisiones,
        "fuera_de_convencion": largos,
        "familias": dict(sorted(familias.items(), key=lambda x: -x[1])),
    }
