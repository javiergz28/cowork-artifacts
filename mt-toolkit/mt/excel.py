"""Lector genérico de reportes de ML.
import warnings
warnings.filterwarnings("ignore", module="openpyxl")

Resuelve dos trampas que aparecen en casi todos los exports:
  1. El encabezado real no está en la fila 1 (hay títulos y notas arriba).
  2. Hay encabezados duplicados ('Unidades' x3, 'Estado' x2, 'Forma de entrega'
     x2) que solo se distinguen por el grupo de la fila de arriba
     ('Ventas', 'Envíos', 'Devoluciones'...).
"""
import openpyxl
from .util import norm_txt


def _leer_filas(path, hoja=None):
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb[hoja] if hoja else wb.worksheets[0]
    filas = [list(r) for r in ws.iter_rows(values_only=True)]
    wb.close()
    return ws.title, filas


def hojas(path):
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    nombres = wb.sheetnames
    wb.close()
    return nombres


def _densidad(fila):
    return sum(1 for c in fila if c is not None and str(c).strip() != "")


def detectar_encabezado(filas, max_scan=15):
    """La fila de encabezado es la primera con más celdas no vacías que la
    siguiente-mejor entre las primeras `max_scan`, y con al menos 4 columnas.
    Devuelve (idx_encabezado, idx_grupo|None).
    """
    mejor, mejor_d = 0, -1
    for i, f in enumerate(filas[:max_scan]):
        d = _densidad(f)
        if d > mejor_d:
            mejor, mejor_d = i, d
    grupo = None
    if mejor > 0:
        prev = filas[mejor - 1]
        # fila de grupo: pocas celdas, todas texto, alineadas a columnas del header
        dp = _densidad(prev)
        if 0 < dp < mejor_d / 2:
            grupo = mejor - 1
    return mejor, grupo


def columnas(filas, i_hdr, i_grp=None):
    """Nombres únicos: 'grupo::columna' cuando hay grupo, con sufijo _2, _3 si
    aún así se repiten. Devuelve lista alineada al ancho del encabezado."""
    hdr = filas[i_hdr]
    grp = filas[i_grp] if i_grp is not None else [None] * len(hdr)
    grp = list(grp) + [None] * (len(hdr) - len(grp))
    actual, gg = None, []
    for g in grp:
        if g is not None and str(g).strip():
            actual = str(g).strip()
        gg.append(actual)
    nombres, vistos = [], {}
    for j, h in enumerate(hdr):
        base = str(h).strip() if h is not None and str(h).strip() else f"col_{j}"
        nom = f"{gg[j]}::{base}" if gg[j] else base
        if nom in vistos:
            vistos[nom] += 1
            nom = f"{nom}_{vistos[nom]}"
        else:
            vistos[nom] = 1
        nombres.append(nom)
    return nombres


def cargar(path, hoja=None):
    """Devuelve (titulo_hoja, nombres_columna, filas_datos, filas_crudas)."""
    titulo, filas = _leer_filas(path, hoja)
    i_hdr, i_grp = detectar_encabezado(filas)
    nombres = columnas(filas, i_hdr, i_grp)
    datos = []
    for f in filas[i_hdr + 1:]:
        if _densidad(f) == 0:
            continue
        f = list(f) + [None] * (len(nombres) - len(f))
        datos.append(dict(zip(nombres, f[:len(nombres)])))
    return titulo, nombres, datos, filas


def buscar_col(nombres, *terminos, grupo=None):
    """Primera columna cuyo nombre contiene todos los términos (sin tildes,
    case-insensitive) y, si se pide, pertenece al grupo indicado."""
    ts = [norm_txt(t) for t in terminos]
    g = norm_txt(grupo) if grupo else None
    for n in nombres:
        nn = norm_txt(n)
        if g and not nn.startswith(g + "::"):
            continue
        if all(t in nn for t in ts):
            return n
    return None
